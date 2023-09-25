# 엑셀파일을 기준으로 데이터를 생성하거나, 테이블을 생성하는 작업

########## import list ##############################################################################################################################
##### Library import 
from PyQt5.QtWidgets import *      # PyQt5 GUI
import logging                      # 로그
import sys                          # 시스템 정보
import openpyxl                     # 엑셀 



########## function #################################################################################################################################

### 엑셀데이터 생성 형태 ▶ excel_obj = [ [타이틀 list들], [수입 list들], [지출 list들], [인건비(지출) list들] ]
### 업로드 form : 급여대장
'''
    @param self
    @return excel_obj(map)
'''
def make_excel_data(self) -> dict :
    try:
        ### 무조건 급여대장(tab_i = 2) 폼에서 한다는 전제로 수행. 추후에는 없애고 요소 명칭 다 변경해야 함.
        wb: openpyxl.Workbook = openpyxl.load_workbook(self.file_path.toPlainText())
        sheet: openpyxl.Workbook.worksheets = wb[wb._sheets[0].title]
        max_col_cnt: int = sheet.max_column

        excel_obj: dict = { ### 이 안에 객체가 4개 들어가야 하므로, 각각의 타이틀, 수입, 지출, 인건비지출 객체리스트를 별도로 생성해야 함.
            'title_list': [],             # 타이틀
            'income_list': [],            # 수입
            'expense_list': [],           # 지출
            'personnel_expense_list': []  # 인건비(지출)
        }

        excel_list: list = [] #다운로드 기능 수행 시 이용할 변수#

        for rows in sheet.iter_rows():
            '''한 행의 데이터를 담을 딕셔너리 자료형 - 엑셀 항목 기준'''
            data_list = [] ## 한 행을 생성하는 list

            for i in range(0, max_col_cnt):
                cell: str = rows[i]

                if str(cell.value) == 'None': input_data: str = ''
                else:                         input_data: str = str(cell.value).replace(' 00:00:00', '') # 거래일자 시분초 제거

                data_list.insert(i, input_data) # list 형태로 삽입(타이틀, 수입, 지출, 인건비(지출))
            
            ### 구분하여 객체별 처리 필요
            if   data_list[0] == '구분':                               excel_obj['title_list'].append(data_list)
            elif data_list[0] == '수입':                               excel_obj['income_list'].append(data_list)
            elif data_list[0] == '지출' and data_list[12] != '인건비': excel_obj['expense_list'].append(data_list)
            elif data_list[0] == '지출' and data_list[12] == '인건비': excel_obj['personnel_expense_list'].append(data_list)

            excel_list.insert(i, data_list) #다운로드 기능 수행 시 이용할 변수#

        #나중에 다운로드하기 위해 미리 대입해둠#
        self.excel_list: list = excel_list

        return excel_obj
    except Exception as e:
        logging.error('엑셀 데이터 생성 실패 : ', str(e))
        sys.exit()


# 엑셀 테이블 생성 : map 데이터 이용
# 엑셀데이터 생성 형태 ▶ excel_obj = [ [타이틀 list들], [수입 list들], [지출 list들], [인건비(지출) list들] ]   
''' 
    @param self, excel_obj(map)
'''
def make_table(self, excel_obj) -> None :
    try:
        ### 급여대장 폼으로 기준을 하여 생성중
        wb: openpyxl = openpyxl.load_workbook(self.file_path.toPlainText())
        excel_tb: openpyxl = self.excel_tb
        status_tb: openpyxl = self.status_tb
        sheet: openpyxl = wb[wb._sheets[0].title]
        max_col_cnt: int = sheet.max_column

        ### 데이터 변수 설정
        title_list: list = excel_obj['title_list']
        income_list: list = excel_obj['income_list']
        expense_list: list = excel_obj['expense_list']
        personnel_expense_list: list = excel_obj['personnel_expense_list']


        ##### 전체 데이터 리스트 생성(조작) ############################################################################
        total_list: list = []

        for a in range(0, len(income_list)): total_list.append(income_list[a]) # 수입 삽입
        for b in range(0, len(expense_list)): total_list.append(expense_list[b]) # 지출 삽입
        for c in range(0, len(personnel_expense_list)): total_list.append(personnel_expense_list[c]) # 인건비(지출) 삽입
        max_row_cnt: int = len(total_list) # 데이터 행 수 삽입
        ##### End 전체 데이터 리스트 생성(조작) #########################################################################

        ### 테이블 세팅
        excel_tb.setColumnCount(max_col_cnt)
        excel_tb.setHorizontalHeaderLabels(title_list[0]) # 타이틀 설정
        excel_tb.setRowCount(max_row_cnt)
        status_tb.setRowCount(max_row_cnt)

        ### 테이블 내 엑셀데이터 표현
        #### 0. max row count idx++
        for i in range(0, max_row_cnt) : 
            data: str = total_list[i]

            for j in range(0, len(data)): excel_tb.setItem(i, j, QTableWidgetItem(data[j]))

        # 상태 테이블 기본설정
        for i in range(0, max_row_cnt): status_tb.setItem(i, 0, QTableWidgetItem('Fail'))
            

        # 마무리 되어서 실무자에게 화면을 보여줄 때 이용할 것.
        # 근데 현재 사용자 입장에서는 이 기능이 필요없을 것 같다....
        # gui.alert('엑셀 데이터를 작업 순서에 맞춰 화면에 표시합니다. \n(이 창은 5초 뒤에 자동으로 닫으므로 건드리지 않습니다.)')
        # time.sleep(5)
    except Exception as e:
        logging.error('엑셀 테이블 생성 실패 : ', str(e))
        sys.exit()