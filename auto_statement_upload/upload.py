# 주제1 : 마음손 전표정보 업로드 자동화(일반전표 -> 인건비(지출) 전표)

########## MEMO ##############################################################################
#1# 파이썬의 명명규칙을 따라서 작명하였음. 규칙을 준수할 것.
#  [ex 1] 변수 및 함수 이름: my_variable, calculate_sum()
#  [ex 2] 클래스 이름: MyClass, MyException

#2# 파이썬 어노테이션(힌트)를 사용하였음. (mypy)
#  [ex] file_path --->> file_path: str




########## import list ##############################################################################
##### Library import 
import os                          # 운영체제 정보
import pyautogui as gui            # 운영체제 제어
import sys                         # 시스템 정보
from PyQt5.QtWidgets import *      # PyQt5 GUI
from PyQt5.QtCore import *
from PyQt5 import uic              # .ui 파일 호출
from PyQt5.QAxContainer import *
from PyQt5.QtGui import *
import logging                     # 로그
import openpyxl                    # 엑셀
import time
from PyQt5.QtCore import *
import keyboard
import signal
import psutil
from pynput.keyboard import Key, Listener

##### Module import 
from module import auto_save             # 결의서/전표정보 자동업로드 및 저장기능 수행
from module import make_excel_data_table # 엑셀 데이터 생성기능 수행
from module import check_data            # input에 작성한 데이터 체크
from module import state
from module import find_and_click
import module.xls_to_xlsx as xls_to_xlsx # 엑셀 확장자 변경




########### 전역처리 ########################################################################################
# 파일경로
main_ui: uic = uic.loadUiType(os.path.dirname(__file__) + os.sep + 'upload_form.ui')[0]

# 공통이미지경로
img_dir_path: os = os.path.dirname(__file__) + os.sep + 'img' + os.sep

# 로그 설정
global_logger: logging = logging.getLogger()  # 전역 로그 (DEBUG 전용)
applogger: logging = logging.getLogger("app") # app 로그 (ERROR 전용)

global_logger.setLevel(logging.ERROR) # DEBUG
applogger.setLevel(logging.DEBUG)     # ERROR

formatter: logging = logging.Formatter(u'%(asctime)s [%(levelname)8s] %(message)s')

# 로그 핸들러 
stream_hander: logging = logging.StreamHandler()
stream_hander.setFormatter(formatter)

applogger.addHandler(stream_hander)     # ERROR

# 파일 핸들러
file_handler: logging = logging.FileHandler('app.log', encoding='utf-8')
file_handler.setFormatter(formatter)

applogger.addHandler(file_handler)     # ERROR



# 기본 딜레이 설정
gui.PAUSE = 0.3
########### Thread ######################################################################################
class thread_stop_fn(QThread) :
    def __init__(self, parent) :
        # 메인에서 받은 self 인자를 parent로 설정함
        super().__init__(parent)
        self.parent = parent

    def run(self) :
        main_window = self.parent

        while True :
            event = keyboard.read_event()

            if (
                event.event_type == keyboard.KEY_DOWN and 
                event.name       == '0'               and
                keyboard.is_pressed('ctrl')
            ) :
                gui.alert('매크로 프로그램을 강제 종료합니다. \n확인을 누르기 전에 다운로드 버튼을 클릭하세요.')
                download(main_window)
                main_pid = os.getpid()
                os.kill(main_pid, signal.SIGTERM) # 시스템 강제종료

                # 이것도 죽여보자
                if state.process == None : 
                    gui.alert('먼저 W4C 프로그램을 실행하세요.')
                    return False
                
                proccess_pid = state.process['pid']
                os.kill(proccess_pid, signal.SIGTERM)
                

class thread_wait_response_fn(QThread) : 
    def __init__(self, parent) :
        super().__init__(parent)
        self.parent = parent
    
    def run(self) :
        # 소스 보관중 언제 또 쓰일지 모름
        if state.process == None : 
            gui.alert('먼저 W4C 프로그램을 실행하세요.')
            return False

        process = psutil.Process(state.process['pid'])

        while True :
            if (
                find_and_click.find_down_img_flag('응답대기2.png') or
                find_and_click.find_down_img_flag('응답대기(비활성화_편집2).png') or
                process.pid != state.process['pid']
            ) :
                current_time = time.localtime()
                gui.alert(f'W4C 프로그램이 응답대기 현상으로 인하여 더 이상 매크로 작업이 불가합니다. \n확인을 누르면 강제종료를 수행합니다. \n(종료시각 : {current_time.tm_hour} : {current_time.tm_min} : {current_time.tm_sec})')
                main_pid = os.getpid()
                os.kill(main_pid, signal.SIGTERM)
                break

# class thread_inner_error_msgbox_fn(QThread) :
#     def __init__(self, parent) :
#         super().__init__(parent)
#         self.parent = parent

#     def run(self) :
#         while True :
#             if find_and_click.find_img_flag('내부서버오류입니다.png') : 
#                 gui.press('enter')
#                 self.parent.inner_error_flag = True
########### class function ##############################################################################
class window__base__setting(QMainWindow, main_ui):
    def __init__(self) :
        super().__init__()

        # 버튼 기능 연결
        self.set_ui()
        self.find_btn.clicked.connect(self.find_fn)                                 # 첨부 엑셀파일
        self.find_payroll_year_img_btn.clicked.connect(self.find_year_img_fn)       # 첨부 회계연도 이미지
        self.start_btn.clicked.connect(self.start_fn)                               # 시작
        self.download_btn.clicked.connect(self.download)                            # 엑셀 다운로드
        self.excel_list = None # make excel data 수행할 때 삽입할 객체
        
        # 스레드 설정 : 강제종료 --> key press event로 처리했으므로 굳이 안해도 됨.
        self.thread_stop_fn = thread_stop_fn(self)
        self.thread_stop_fn.start()

        # 첫 실행때만 프로세스 정보를 status에 담는다.
        while True :
            for p in psutil.process_iter(attrs=['pid', 'name', 'status', 'cmdline']) :
                if p.info['name'].lower() == 'xplatform.exe' :
                    state.process = p.info
                    break
                else :
                    continue
            break

        # 스레드 설정 : 응답대기 현상 발생 시 매크로 강제종료
        self.thread_wait_response_fn = thread_wait_response_fn(self)
        self.thread_wait_response_fn.start()

        # 스레드 설정 : '내부오류입니다' 메세지 발견 시 엔터
        # self.thread_inner_error_msgbox_fn = thread_inner_error_msgbox_fn(self)
        # self.thread_inner_error_msgbox_fn.start()
        # 내부오류입니다와 관련된 변수 설정
        # self.inner_error_flag = False

    # ui 세팅
    def set_ui(self) -> None : self.setupUi(self)

    #1 파일 업로드
    def find_fn(self) -> None :
        try:
            file_path: QFileDialog = QFileDialog.getOpenFileName(self)
            file_nm: os = os.path.basename(file_path[0])

            change_file_path: str = self.file_path

            if ('.xlsx' in file_nm) or ('.xls' in file_nm):
                if 'xls' == file_nm.split('.')[1]: 
                    xls_to_xlsx.xls_to_xlsx(file_path[0]) # 파일변환 작업
                    change_file_path.setText(file_path[0] + 'x')
                else: change_file_path.setText(file_path[0])

                # 엑셀 테이블 생성
                state.excel_obj = make_excel_data_table.make_excel_data(self)
                make_excel_data_table.make_table(self, state.excel_obj)
            else: gui.alert('xlsx 또는 xls 확장자만 허용합니다.')
        except Exception as e: 
            gui.alert('파일업로드 과정에서 오류가 발생했습니다. \n관리자 확인이 필요합니다.')
            logging.error(str(e))


    # 전표등록 자동업로드 시작
    def start_fn(self) :
        if gui.confirm('전표정보 자동업로드 업무를 실행하시겠습니까?') == 'OK':
            starting(self)

            # 확인사항 조건이 맞으면 자동업로드 시작
            if check_data.select_and_return_result(self) : start_auto(self)
        else: 
            gui.alert('전표정보 자동업로드 실행을 취소합니다.')
            self.stop_fn


    #3 자동업로드 중지
    def stop_fn(self) :
        gui.alert('자동화 업무를 중단합니다.')
        time.sleep(0.3)
        ending(self)
        time.sleep(0.3)
        sys.exit()


    #4 사업명 이미지 업로드
    def find_project_img_fn(self) -> None :
        try:
            file_path: str = QFileDialog.getOpenFileName(self)

            ## 업로드하는 이미지명에 '인건비' 포함여부 확인하여 file_img_path 설정하기
            if   '인건비' in file_path[0]:     change_file_path: str = self.file_payroll_project_img_path
            elif '인건비' not in file_path[0]: change_file_path: str = self.file_project_img_path

            change_file_path.setText(file_path[0]) if ('.png' in file_path[0]) else gui.alert('png 확장자 이미지만 허용합니다.')
        except Exception as e:
            gui.alert('사업명 이미지 파일업로드 과정에서 오류가 발생했습니다. \n관리자 확인이 필요합니다.')
            logging.error(str(e))


    # 회계연도 이미지 업로드
    def find_year_img_fn(self) -> None :
        try:
            file_path: str = QFileDialog.getOpenFileName(self)
            change_file_path: str = self.file_payroll_year_img_path
            change_file_path.setText(file_path[0]) if ('.png' in file_path[0]) else gui.alert('png 확장자 이미지만 허용합니다.')
        except Exception as e:
            gui.alert('회계연도 이미지 파일업로드 과정에서 오류가 발생했습니다. \n관리자 확인이 필요합니다.')
            logging.error(str(e))


    # 엑셀 생성 + 엑셀 다운로드
    def download(self) -> None :
        excel_tb  = self.excel_tb
        status_tb = self.status_tb
        # excel_list = self.excel_list #실제로 담겨져있는 엑셀데이터#
        excel_obj = state.excel_obj #실제로 담겨져있는 엑셀객체#
        custom_list = []

        #엑셀 행 수가 0 이상인 경우에만 수행#
        if excel_tb.rowCount() > 0:
            if gui.confirm('작업한 결과를 저장하시겠습니까?') == 'OK':
                try:
                    #Workbook 생성#
                    wb: openpyxl.Workbook = openpyxl.Workbook()
                    #저장경로 추출하기 위한 요소 조회#
                    file_path: str = self.file_path.toPlainText()
                    file_rsplit: str = file_path.rsplit('/')
                    length: int = len(file_rsplit)
                        # 파일명 추출
                    file_nm: str = file_rsplit[length-1]
                    file_rsplit.pop()
                        # 저장경로 생성
                    save_path: str = '/'.join(file_rsplit) + '/'
                        # 저장파일명 생성
                    save_file_nm: str = '[결과]' + file_nm


                    #엑셀내용 생성 시작#
                        # 현재 워크시트 선택
                    ws: openpyxl = wb.active

                        # 첫행은 무조건 타이틀 삽입
                    ws.append(excel_obj['title_list'][0])

                    # custom_list에 수입, 지출, 인건비(지출) 모두 담기
                    row_cnt = self.excel_tb.rowCount()
                    col_cnt = self.excel_tb.columnCount()

                    for i in range(0, row_cnt) :
                        row_list = []

                        for j in range(0, col_cnt) :
                            cell_data = excel_tb.item(i,j).text()
                            row_list.append(cell_data)
                        custom_list.insert(i, row_list)

                        # append()를 이용하여 list 자체를 하나의 row로 채운다
                    for idx in range(1, len(custom_list) + 1):
                        # 행에 데이터 자체를 붙임
                        ws.append(custom_list[idx - 1])
                        # 셀 범위 설정
                        cell_range: str = f'A{idx + 1}:{openpyxl.utils.get_column_letter(ws.max_column)}{idx + 1}'
                        # 채우기 색상 설정
                        fill: openpyxl = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type="solid")
                        # 등록하는 전표 상태값이 'Success'이면 넘어가고, 'Fail' 이면 배경색 칠하기
                        # status_tb와 excel_tb의 list size() 값이 다르므로 idx-1 처리
                        if status_tb.item(idx - 1, 0).text() == 'Fail':
                            for row in ws[cell_range]:
                                for cell in row: 
                                    cell.fill = fill


                    # 엑셀 저장(마무리)
                    wb.save(save_path + save_file_nm)
                except Exception as e:
                    gui.alert('첨부파일 엑셀작업 결과를 엑셀로 생성하는 과정에서 문제가 발생했습니다.')

########## function ###################################################################################
# 자동화 실행 >> 전표정보
def start_auto(self) -> None :
    # Action
    applogger.debug('----- 매크로 START -----')
    # excel_obj: object = make_excel_data_table.make_excel_data(self)
    # make_excel_data_table.make_table(self, excel_obj)
    excel_obj = state.excel_obj

    # Active
    w4c_window = None

    if len(gui.getWindowsWithTitle('사회복지시설정보시스템(1W)')) > 0 :
        w4c_window = gui.getWindowsWithTitle('사회복지시설정보시스템(1W)')[0]
    else :
        gui.alert('매크로 프로그램을 사용할 대상 프로그램을 실행하세요.')
        return False
    
    if w4c_window.isActive == False: w4c_window.activate()

    auto_save.auto_save(self, excel_obj)
    ending(self)
    applogger.debug('----- 매크로 END -----')


#6# '실행중'으로 상태변경
def starting(self) -> None :
    self.status_text.setText('실행중')
    self.status_text.setStyleSheet('color: red')

#7# '종료'으로 상태변경
def ending(self) -> None :
    self.status_text.setText('종료')
    self.status_text.setStyleSheet('Color: black')


# 엑셀 생성 + 엑셀 다운로드
def download(self) -> None :
    excel_tb  = self.excel_tb
    status_tb = self.status_tb
    # excel_list = self.excel_list #실제로 담겨져있는 엑셀데이터#
    excel_obj = state.excel_obj #실제로 담겨져있는 엑셀객체#
    custom_list = []

    #엑셀 행 수가 0 이상인 경우에만 수행#
    if excel_tb.rowCount() > 0:
        if gui.confirm('작업한 결과를 저장하시겠습니까?') == 'OK':
            try:
                #Workbook 생성#
                wb: openpyxl.Workbook = openpyxl.Workbook()
                #저장경로 추출하기 위한 요소 조회#
                file_path: str = self.file_path.toPlainText()
                file_rsplit: str = file_path.rsplit('/')
                length: int = len(file_rsplit)
                    # 파일명 추출
                file_nm: str = file_rsplit[length-1]
                file_rsplit.pop()
                    # 저장경로 생성
                save_path: str = '/'.join(file_rsplit) + '/'
                    # 저장파일명 생성
                save_file_nm: str = '[결과]' + file_nm


                #엑셀내용 생성 시작#
                    # 현재 워크시트 선택
                ws: openpyxl = wb.active

                    # 첫행은 무조건 타이틀 삽입
                ws.append(excel_obj['title_list'][0])

                # custom_list에 수입, 지출, 인건비(지출) 모두 담기
                row_cnt = self.excel_tb.rowCount()
                col_cnt = self.excel_tb.columnCount()

                for i in range(0, row_cnt) :
                    row_list = []

                    for j in range(0, col_cnt) :
                        cell_data = excel_tb.item(i,j).text()
                        row_list.append(cell_data)
                    custom_list.insert(i, row_list)

                    # append()를 이용하여 list 자체를 하나의 row로 채운다
                for idx in range(1, len(custom_list) + 1):
                    # 행에 데이터 자체를 붙임
                    ws.append(custom_list[idx - 1])
                    
                    # 셀 범위 설정
                    cell_range: str = f'A{idx + 1}:{openpyxl.utils.get_column_letter(ws.max_column)}{idx + 1}'
                    # 채우기 색상 설정
                    fill: openpyxl = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type="solid")
                    # 등록하는 전표 상태값이 'Success'이면 넘어가고, 'Fail' 이면 배경색 칠하기
                    # status_tb와 excel_tb의 list size() 값이 다르므로 idx-1 처리
                    if status_tb.item(idx - 1, 0).text() == 'Fail':
                        for row in ws[cell_range]:
                            for cell in row: 
                                cell.fill = fill


                # 엑셀 저장(마무리)
                wb.save(save_path + save_file_nm)
            except Exception as e:
                gui.alert('첨부파일 엑셀작업 결과를 엑셀로 생성하는 과정에서 문제가 발생했습니다.')






# 메인함수
def main() :
    app = QApplication(sys.argv)
    window = window__base__setting()
    window.show()
    # QApplication 창을 오픈하는 것이므로 마지막 실행
    app.exec_()


########## Start Program(PyQt5 Designer) ###################################################################################
if __name__ == '__main__' :
    main()